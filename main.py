from email.mime import message
import psycopg2
import telebot
import datetime
import config
from telebot import types
from openpyxl import Workbook, load_workbook
import time

bot = telebot.TeleBot(config.CONFIG['token'])

db_con = psycopg2.connect("dbname=d88e8nu9n2ngq1 user=pfonljqshtvuzv host=ec2-54-228-97-176.eu-west-1.compute.amazonaws.com password=09d7f92ad0900c867d6233570339398aba1d5677c711ef5ded2214b056d0eef0 port=5432", sslmode="require")
db_cur = db_con.cursor()


class Person:
    ind = '0'
    team = 'default'
    user_info = ()

    def __init__(self, id_person=ind, team_person=team):
        self.ind = id_person
        self.team = str(team_person)
        self.user_info = (self.team, self.ind)

    def write_data(self):
        db_cur.execute("SELECT id_user FROM users")
        x = db_cur.fetchall()
        db_con.commit()

        data = [a[0] for a in x]
        if self.ind in data:
            db_cur.execute(
                "UPDATE users SET team_user = %s WHERE id_user = %s", self.user_info)
            db_con.commit()
        else:
            db_cur.execute(
                "INSERT INTO users(team_user, id_user) VALUES (%s, %s);", self.user_info)
            db_con.commit()

    def get_user_team(self, id_user):
        db_cur.execute(
            f"SELECT team_user FROM users WHERE id_user = {id_user};")
        return db_cur.fetchone()


def reply_get_user_info(message, mode):
    time.sleep(1)
    if mode == 1:
        rek_kb = types.ReplyKeyboardMarkup(
            resize_keyboard=True, one_time_keyboard=True)
        btn_reply = types.KeyboardButton('Выбрать снова команду')
        btn_main = types.KeyboardButton('Вернуться в главное меню')
        rek_kb.add(btn_reply, btn_main)

        bot.send_message(message.chat.id, 'Нажми на кнопку',
                         reply_markup=rek_kb)
    else:
        main_info(message)


def main_info(message):
    first_kb = types.InlineKeyboardMarkup()
    btn_all = types.InlineKeyboardButton(
        text='Все команды', callback_data='all')
    btn_single = types.InlineKeyboardButton(
        text='Любимая команда', callback_data='single')
    first_kb.add(btn_all, btn_single)

    bot.send_message(message.chat.id, "Мы переместились в главное меню😎")
    bot.send_message(message.chat.id, "1. Ты можешь узнать расписание команды выбирая её из списка всех команд\n2. Ты можешь узнать расписание своей любимой команды\nВыбор за тобой🤫", reply_markup=first_kb)


def true_time(time):
    time_shd = time.split(':')
    time_shd = [int(time_shd[i]) for i in range(len(time.split(':')))]
    time_shd[0] += 3
    return time_shd


def print_game(team, message, mode):
    wb = load_workbook("TrueShd.xlsx")
    sheet = wb[str(team)]

    date_today = str(datetime.date.today()).split('-')
    time_today = str(datetime.datetime.today())[11:16].split(':')
    for i in range(sheet.max_row):
        date_shd = (str(sheet["D" + str(i + 1)].value)).split(":")
        if date_today[0] == date_shd[0] and date_today[1] == date_shd[1]:
            if date_today[2] == date_shd[2]:
                time_shd = true_time(sheet['C' + str(i + 1)].value)
                if int(time_today[0]) <= time_shd[0] and int(time_today[1]) < time_shd[1]:
                    bot.send_message(
                        message.chat.id, f"Следующая игра команды {config.TEAM[team]}\nДата: {sheet['A' + str(i + 1)].value}\nВремя: {sheet['C' + str(i + 1)].value}\nПротив команды {sheet['B' + str(i + 1)].value}")
                    reply_get_user_info(message, mode)
                    break
                else:
                    continue
            elif int(date_today[2]) < int(date_shd[2]):
                bot.send_message(
                    message.chat.id, f"Следующая игра команды {config.TEAM[team]}\nДата: {sheet['A' + str(i + 1)].value}\nВремя: {sheet['C' + str(i + 1)].value}\nПротив команды {sheet['B' + str(i + 1)].value}")
                reply_get_user_info(message, mode)
                break
            else:
                continue
        else:
            continue


def first_choos_user(message):
    global choos_id
    choos_id = 1
    bot.send_message(
        message.chat.id, "Давай сначала ты выберешь любимую команду (ты её сможешь поменять в любой момент)")
    bot.send_message(
        message.chat.id, "Выбери конференцию", reply_markup=m_inl)


@bot.message_handler(commands=['start'])
def start_user_info(message):
    global id_user
    id_user = message.from_user.id
    bot.send_message(message.chat.id, 'Привет, Я Бот "Shedule NBA"')
    first_choos_user(message)


@bot.message_handler(content_types=['text'])
def eror_message(message):
    global choos_id, user
    if message.text == 'Выбрать снова команду':
        bot.send_message(
            message.chat.id, "Выбери конференцию", reply_markup=m_inl)
    elif message.text == 'Вернуться в главное меню':
        main_info(message)
    elif message.text == 'Узнать время':
        id_user = message.from_user.id
        u_team = user.get_user_team(id_user)
        print_game(u_team[0].strip(), message, 2)
    elif message.text == 'Поменять команду':
        choos_id = 2
        bot.send_message(
            message.chat.id, "Выбери конференцию", reply_markup=m_inl)
    else:
        bot.send_message(message.chat.id, "Прости я не понимаю твою команду😔")
        main_info(message)


@bot.callback_query_handler(func=lambda call: True)
def answer(call):
    global choos_id, user, id_user
    del_mes = call.message.message_id - 1
    if call.data == 'all':
        choos_id = 0
        bot.delete_message(chat_id=call.message.chat.id, message_id=del_mes)
        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                              text='Выбери конференцию', reply_markup=m_inl)

    elif call.data == 'single':
        choos_id = 2
        bot.delete_message(chat_id=call.message.chat.id, message_id=del_mes)
        rek_kb = types.ReplyKeyboardMarkup(
            resize_keyboard=True, one_time_keyboard=True)
        btn_reply = types.KeyboardButton('Узнать время')
        btn_main = types.KeyboardButton('Поменять команду')
        rek_kb.add(btn_reply, btn_main)
        bot.delete_message(chat_id=call.message.chat.id,
                           message_id=call.message.message_id)
        bot.send_message(chat_id=call.message.chat.id,
                         text='Выбери и нажми на кнопку', reply_markup=rek_kb)

    elif call.data == 'east':
        meast_inl = types.InlineKeyboardMarkup()
        btn_BOS = types.InlineKeyboardButton(
            text='Бостон Селтикс', callback_data='BOS')
        btn_NYK = types.InlineKeyboardButton(
            text='Нью-Йорк Никс', callback_data='NYK')
        btn_BRK = types.InlineKeyboardButton(
            text='Бруклин Нетс', callback_data='BRK')
        btn_PHI = types.InlineKeyboardButton(
            text='Филадельфия 76 Сиксерс', callback_data='PHI')
        btn_TOR = types.InlineKeyboardButton(
            text='Торонто Рапторз', callback_data='TOR')
        btn_ATL = types.InlineKeyboardButton(
            text='Атланта Хоукс', callback_data='ATL')
        btn_CHO = types.InlineKeyboardButton(
            text='Шарлотт Хорнетс', callback_data='CHO')
        btn_MIA = types.InlineKeyboardButton(
            text='Майами Хит', callback_data='MIA')
        btn_ORL = types.InlineKeyboardButton(
            text='Орладно Мэджик', callback_data='ORL')
        btn_WAS = types.InlineKeyboardButton(
            text='Вашингтон Уизардс', callback_data='WAS')
        btn_CHI = types.InlineKeyboardButton(
            text='Чикаго Буллз', callback_data='CHI')
        btn_CLE = types.InlineKeyboardButton(
            text='Кливленд Кавальерс', callback_data='CLE')
        btn_DET = types.InlineKeyboardButton(
            text='Детроит Пистонс', callback_data='DET')
        btn_IND = types.InlineKeyboardButton(
            text='Индиана Пэйсерс', callback_data='IND')
        btn_MIL = types.InlineKeyboardButton(
            text='Милуоки Бакс', callback_data='MIL')
        meast_inl.add(
            btn_BOS, btn_NYK, btn_BRK, btn_PHI, btn_TOR,
            btn_ATL, btn_CHO, btn_MIA, btn_ORL, btn_WAS,
            btn_CHI, btn_CLE, btn_DET, btn_IND, btn_MIL)

        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                              text='Выбери команду', reply_markup=meast_inl)

    elif call.data == 'west':
        mwest_inl = types.InlineKeyboardMarkup()
        btn_POR = types.InlineKeyboardButton(
            text='Портленд Трейл Блейзерс', callback_data='POR')
        btn_MIN = types.InlineKeyboardButton(
            text='Миннесота Тимбервулз', callback_data='MIN')
        btn_OKC = types.InlineKeyboardButton(
            text='Оклахома-Сити Тандер', callback_data='OKC')
        btn_DEN = types.InlineKeyboardButton(
            text='Денвер Наггетс', callback_data='DEN')
        btn_UTA = types.InlineKeyboardButton(
            text='Юта Джаз', callback_data='UTA')
        btn_DAL = types.InlineKeyboardButton(
            text='Даллас Маверикс', callback_data='DAL')
        btn_HOU = types.InlineKeyboardButton(
            text='Хьюстон Рокетс', callback_data='HOU')
        btn_MEM = types.InlineKeyboardButton(
            text='Мемфис Гриззлис', callback_data='MEM')
        btn_NOP = types.InlineKeyboardButton(
            text='Нью-Орлеан Пеликанс', callback_data='NOP')
        btn_SAS = types.InlineKeyboardButton(
            text='Сан-Антонио Спёрс', callback_data='SAS')
        btn_GSW = types.InlineKeyboardButton(
            text='Голден Стэйт Уорриорз', callback_data='GSW')
        btn_LAC = types.InlineKeyboardButton(
            text='ЛА Клипперс', callback_data='LAC')
        btn_LAL = types.InlineKeyboardButton(
            text='ЛА Лейкерс', callback_data='LAL')
        btn_PHO = types.InlineKeyboardButton(
            text='Финикс Санз', callback_data='PHO')
        btn_SAC = types.InlineKeyboardButton(
            text='Сакраменто Кингз', callback_data='SAC')
        mwest_inl.add(
            btn_POR, btn_MIN, btn_OKC, btn_DEN, btn_UTA,
            btn_DAL, btn_HOU, btn_MEM, btn_NOP, btn_SAS,
            btn_GSW, btn_LAC, btn_LAL, btn_PHO, btn_SAC)

        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                              text='Выбери команду', reply_markup=mwest_inl)

    else:
        if choos_id == 1 or choos_id == 2:
            team_user = call.data
            user = Person(id_user, team_user)
            user.write_data()
            if choos_id == 1:
                bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                      text="Отлично, Ты теперь в любой момент можешь посмотреть расписание своей любимой команды", reply_markup=None)
            else:
                bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                      text="Ты поменял команду", reply_markup=None)
            time.sleep(1)
            main_info(call.message)
        else:
            bot.delete_message(chat_id=call.message.chat.id,
                               message_id=call.message.message_id)
            print_game(call.data, call.message, 1)


m_inl = types.InlineKeyboardMarkup()
btn_east = types.InlineKeyboardButton(text='Восточная', callback_data='east')
btn_west = types.InlineKeyboardButton(text='Западная', callback_data='west')
m_inl.add(btn_east, btn_west)

user = Person()
choos_id = 0
id_user = 0

bot.polling(none_stop=True, interval=0)
